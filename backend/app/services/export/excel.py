"""Excel export service for MOZG Analytics."""

import io
import uuid
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.reports.sales import SalesReportService, SalesDataPoint, VenueSales
from app.services.reports.menu import (
    MenuAnalysisService,
    ProductABC,
    ProductMargin,
    GoListItem,
)


class ExcelExportService:
    """Service for exporting reports to Excel format."""

    # Styles
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    ABC_A_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ABC_B_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ABC_C_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self, db: AsyncSession):
        self.db = db
        self.sales_service = SalesReportService(db)
        self.menu_service = MenuAnalysisService(db)

    def _apply_header_style(self, cell):
        """Apply header style to a cell."""
        cell.fill = self.HEADER_FILL
        cell.font = self.HEADER_FONT
        cell.alignment = self.HEADER_ALIGNMENT
        cell.border = self.BORDER

    def _apply_cell_style(self, cell, alignment: str = "left"):
        """Apply standard cell style."""
        cell.border = self.BORDER
        cell.alignment = Alignment(horizontal=alignment, vertical="center")

    def _auto_column_width(self, ws, min_width: int = 10, max_width: int = 50):
        """Auto-adjust column widths based on content."""
        for column_cells in ws.columns:
            max_length = 0
            column = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column].width = adjusted_width

    async def export_sales_summary(
        self,
        venue_ids: List[uuid.UUID],
        date_from: date,
        date_to: date,
        include_daily: bool = True,
        include_hourly: bool = True,
        include_venues: bool = True,
    ) -> bytes:
        """
        Export sales summary report to Excel.

        Args:
            venue_ids: List of venue UUIDs
            date_from: Start date
            date_to: End date
            include_daily: Include daily breakdown sheet
            include_hourly: Include hourly breakdown sheet
            include_venues: Include venue comparison sheet

        Returns:
            Excel file as bytes
        """
        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        await self._write_sales_summary_sheet(ws_summary, venue_ids, date_from, date_to)

        # Daily breakdown
        if include_daily:
            ws_daily = wb.create_sheet("Daily Sales")
            await self._write_daily_sales_sheet(ws_daily, venue_ids, date_from, date_to)

        # Hourly breakdown
        if include_hourly:
            ws_hourly = wb.create_sheet("Hourly Analysis")
            await self._write_hourly_sales_sheet(ws_hourly, venue_ids, date_from, date_to)

        # Venue comparison
        if include_venues:
            ws_venues = wb.create_sheet("By Venue")
            await self._write_venue_sales_sheet(ws_venues, venue_ids, date_from, date_to)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def _write_sales_summary_sheet(
        self,
        ws,
        venue_ids: List[uuid.UUID],
        date_from: date,
        date_to: date,
    ):
        """Write sales summary to worksheet."""
        summary = await self.sales_service.get_summary(venue_ids, date_from, date_to)

        # Title
        ws["A1"] = "Sales Summary Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Period: {date_from.strftime('%Y-%m-%d')} - {date_to.strftime('%Y-%m-%d')}"
        ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # Metrics
        metrics = [
            ("Total Revenue", f"{summary.revenue:,.2f}"),
            ("Number of Receipts", f"{summary.receipts_count:,}"),
            ("Average Check", f"{summary.avg_check:,.2f}"),
            ("Total Guests", f"{summary.guests_count:,}"),
            ("Total Items Sold", f"{summary.items_count:,}"),
            ("Items per Receipt", f"{summary.items_per_receipt:.2f}"),
            ("Revenue per Guest", f"{summary.revenue_per_guest:,.2f}"),
            ("Total Discounts", f"{summary.total_discount:,.2f}"),
        ]

        row = 5
        ws[f"A{row}"] = "Metric"
        ws[f"B{row}"] = "Value"
        self._apply_header_style(ws[f"A{row}"])
        self._apply_header_style(ws[f"B{row}"])

        for metric_name, metric_value in metrics:
            row += 1
            ws[f"A{row}"] = metric_name
            ws[f"B{row}"] = metric_value
            self._apply_cell_style(ws[f"A{row}"])
            self._apply_cell_style(ws[f"B{row}"], "right")

        self._auto_column_width(ws)

    async def _write_daily_sales_sheet(
        self,
        ws,
        venue_ids: List[uuid.UUID],
        date_from: date,
        date_to: date,
    ):
        """Write daily sales breakdown to worksheet."""
        daily_data = await self.sales_service.get_daily(venue_ids, date_from, date_to)

        # Headers
        headers = ["Date", "Revenue", "Receipts", "Avg Check", "Guests"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Data
        for row_idx, dp in enumerate(daily_data, 2):
            ws.cell(row=row_idx, column=1, value=dp.date.strftime("%Y-%m-%d"))
            ws.cell(row=row_idx, column=2, value=float(dp.revenue))
            ws.cell(row=row_idx, column=3, value=dp.receipts_count)
            ws.cell(row=row_idx, column=4, value=float(dp.avg_check))
            ws.cell(row=row_idx, column=5, value=dp.guests_count)

            for col in range(1, 6):
                self._apply_cell_style(ws.cell(row=row_idx, column=col))

        self._auto_column_width(ws)

    async def _write_hourly_sales_sheet(
        self,
        ws,
        venue_ids: List[uuid.UUID],
        date_from: date,
        date_to: date,
    ):
        """Write hourly sales analysis to worksheet."""
        hourly_data = await self.sales_service.get_hourly(venue_ids, date_from, date_to)

        # Headers
        headers = ["Hour", "Total Revenue", "Total Receipts", "Avg Revenue per Day"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Data
        for row_idx, h in enumerate(hourly_data, 2):
            ws.cell(row=row_idx, column=1, value=f"{h.hour:02d}:00")
            ws.cell(row=row_idx, column=2, value=float(h.revenue))
            ws.cell(row=row_idx, column=3, value=h.receipts_count)
            ws.cell(row=row_idx, column=4, value=float(h.avg_revenue))

            for col in range(1, 5):
                self._apply_cell_style(ws.cell(row=row_idx, column=col))

        self._auto_column_width(ws)

    async def _write_venue_sales_sheet(
        self,
        ws,
        venue_ids: List[uuid.UUID],
        date_from: date,
        date_to: date,
    ):
        """Write venue comparison to worksheet."""
        venue_data = await self.sales_service.get_by_venue(venue_ids, date_from, date_to)

        # Headers
        headers = ["Venue", "Revenue", "% of Total", "Receipts", "Avg Check", "Guests"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Data
        for row_idx, v in enumerate(venue_data, 2):
            ws.cell(row=row_idx, column=1, value=v.venue_name)
            ws.cell(row=row_idx, column=2, value=float(v.revenue))
            ws.cell(row=row_idx, column=3, value=f"{v.revenue_percent:.1f}%")
            ws.cell(row=row_idx, column=4, value=v.receipts_count)
            ws.cell(row=row_idx, column=5, value=float(v.avg_check))
            ws.cell(row=row_idx, column=6, value=v.guests_count)

            for col in range(1, 7):
                self._apply_cell_style(ws.cell(row=row_idx, column=col))

        self._auto_column_width(ws)

    async def export_abc_analysis(
        self,
        venue_ids: List[uuid.UUID],
        date_from: date,
        date_to: date,
        metric: str = "revenue",
    ) -> bytes:
        """
        Export ABC analysis to Excel.

        Args:
            venue_ids: List of venue UUIDs
            date_from: Start date
            date_to: End date
            metric: Metric for analysis (revenue, profit, quantity)

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "ABC Analysis"

        result = await self.menu_service.abc_analysis(venue_ids, date_from, date_to, metric)

        # Title
        ws["A1"] = f"ABC Analysis by {metric.title()}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Period: {date_from.strftime('%Y-%m-%d')} - {date_to.strftime('%Y-%m-%d')}"

        # Summary
        ws["A4"] = "Category Summary"
        ws["A4"].font = Font(bold=True)

        summary_headers = ["Category", "Products", "Revenue", "Profit", "% of Total"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            self._apply_header_style(cell)

        row = 6
        for cat, data in result.summary.items():
            ws.cell(row=row, column=1, value=cat.value)
            ws.cell(row=row, column=2, value=data["count"])
            ws.cell(row=row, column=3, value=float(data["revenue"]))
            ws.cell(row=row, column=4, value=float(data["profit"]))
            ws.cell(row=row, column=5, value=f"{data['revenue_percent']:.1f}%")

            # Color by category
            if cat.value == "A":
                ws.cell(row=row, column=1).fill = self.ABC_A_FILL
            elif cat.value == "B":
                ws.cell(row=row, column=1).fill = self.ABC_B_FILL
            else:
                ws.cell(row=row, column=1).fill = self.ABC_C_FILL

            for col in range(1, 6):
                self._apply_cell_style(ws.cell(row=row, column=col))
            row += 1

        # Product details
        ws.cell(row=row + 1, column=1, value="Product Details")
        ws.cell(row=row + 1, column=1).font = Font(bold=True)

        headers = [
            "Product",
            "Category",
            "Qty",
            "Revenue",
            "Cost",
            "Profit",
            "Margin %",
            "Revenue %",
            "Cumulative %",
            "ABC",
        ]
        header_row = row + 2
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            self._apply_header_style(cell)

        for idx, p in enumerate(result.products, header_row + 1):
            ws.cell(row=idx, column=1, value=p.product_name)
            ws.cell(row=idx, column=2, value=p.category_name or "")
            ws.cell(row=idx, column=3, value=float(p.quantity))
            ws.cell(row=idx, column=4, value=float(p.revenue))
            ws.cell(row=idx, column=5, value=float(p.cost))
            ws.cell(row=idx, column=6, value=float(p.profit))
            ws.cell(row=idx, column=7, value=f"{p.margin_percent:.1f}%")
            ws.cell(row=idx, column=8, value=f"{p.revenue_percent:.2f}%")
            ws.cell(row=idx, column=9, value=f"{p.cumulative_percent:.2f}%")
            ws.cell(row=idx, column=10, value=p.abc_category.value)

            # Color ABC column
            if p.abc_category.value == "A":
                ws.cell(row=idx, column=10).fill = self.ABC_A_FILL
            elif p.abc_category.value == "B":
                ws.cell(row=idx, column=10).fill = self.ABC_B_FILL
            else:
                ws.cell(row=idx, column=10).fill = self.ABC_C_FILL

            for col in range(1, 11):
                self._apply_cell_style(ws.cell(row=idx, column=col))

        self._auto_column_width(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def export_go_list(
        self,
        venue_ids: List[uuid.UUID],
        date_from: date,
        date_to: date,
        margin_threshold: Optional[Decimal] = None,
    ) -> bytes:
        """
        Export Go-List to Excel.

        Args:
            venue_ids: List of venue UUIDs
            date_from: Start date
            date_to: End date
            margin_threshold: Margin threshold for classification

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Go-List"

        result = await self.menu_service.go_list(
            venue_ids, date_from, date_to, margin_threshold
        )

        # Title
        ws["A1"] = "Go-List Analysis"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Period: {date_from.strftime('%Y-%m-%d')} - {date_to.strftime('%Y-%m-%d')}"

        # Recommendations
        ws["A4"] = "Key Recommendations"
        ws["A4"].font = Font(bold=True)
        for idx, rec in enumerate(result.recommendations, 5):
            ws[f"A{idx}"] = f"• {rec}"

        # Summary
        start_row = 5 + len(result.recommendations) + 1
        ws.cell(row=start_row, column=1, value="Category Summary")
        ws.cell(row=start_row, column=1).font = Font(bold=True)

        summary_headers = ["Category", "Products", "Revenue", "Profit"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=start_row + 1, column=col, value=header)
            self._apply_header_style(cell)

        category_fills = {
            "stars": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
            "workhorses": PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
            "puzzles": PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid"),
            "dogs": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "potential": PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid"),
            "standard": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
        }

        row = start_row + 2
        for cat, data in result.summary.items():
            ws.cell(row=row, column=1, value=cat.value.title())
            ws.cell(row=row, column=2, value=data["count"])
            ws.cell(row=row, column=3, value=float(data["revenue"]))
            ws.cell(row=row, column=4, value=float(data["profit"]))

            if cat.value in category_fills:
                ws.cell(row=row, column=1).fill = category_fills[cat.value]

            for col in range(1, 5):
                self._apply_cell_style(ws.cell(row=row, column=col))
            row += 1

        # Product details
        ws.cell(row=row + 1, column=1, value="Product Details")
        ws.cell(row=row + 1, column=1).font = Font(bold=True)

        headers = [
            "Product",
            "Category",
            "ABC",
            "Margin %",
            "Go-List",
            "Revenue",
            "Profit",
            "Recommendation",
        ]
        header_row = row + 2
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            self._apply_header_style(cell)

        for idx, item in enumerate(result.items, header_row + 1):
            ws.cell(row=idx, column=1, value=item.product_name)
            ws.cell(row=idx, column=2, value=item.category_name or "")
            ws.cell(row=idx, column=3, value=item.abc_category.value)
            ws.cell(row=idx, column=4, value=f"{item.margin_percent:.1f}%")
            ws.cell(row=idx, column=5, value=item.go_list_category.value.title())
            ws.cell(row=idx, column=6, value=float(item.revenue))
            ws.cell(row=idx, column=7, value=float(item.profit))
            ws.cell(row=idx, column=8, value=item.recommendation)

            if item.go_list_category.value in category_fills:
                ws.cell(row=idx, column=5).fill = category_fills[item.go_list_category.value]

            for col in range(1, 9):
                self._apply_cell_style(ws.cell(row=idx, column=col))

        self._auto_column_width(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    async def export_margin_analysis(
        self,
        venue_ids: List[uuid.UUID],
        date_from: date,
        date_to: date,
        min_quantity: int = 1,
    ) -> bytes:
        """
        Export margin analysis to Excel.

        Args:
            venue_ids: List of venue UUIDs
            date_from: Start date
            date_to: End date
            min_quantity: Minimum quantity filter

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Margin Analysis"

        margins = await self.menu_service.margin_analysis(
            venue_ids, date_from, date_to, min_quantity
        )

        # Title
        ws["A1"] = "Product Margin Analysis"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Period: {date_from.strftime('%Y-%m-%d')} - {date_to.strftime('%Y-%m-%d')}"

        # Headers
        headers = [
            "Product",
            "Category",
            "Quantity",
            "Revenue",
            "Cost",
            "Profit",
            "Margin %",
            "Avg Price",
            "Avg Cost",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            self._apply_header_style(cell)

        # Data
        for idx, m in enumerate(margins, 5):
            ws.cell(row=idx, column=1, value=m.product_name)
            ws.cell(row=idx, column=2, value=m.category_name or "")
            ws.cell(row=idx, column=3, value=float(m.quantity))
            ws.cell(row=idx, column=4, value=float(m.revenue))
            ws.cell(row=idx, column=5, value=float(m.cost))
            ws.cell(row=idx, column=6, value=float(m.profit))
            ws.cell(row=idx, column=7, value=f"{m.margin_percent:.1f}%")
            ws.cell(row=idx, column=8, value=float(m.avg_price))
            ws.cell(row=idx, column=9, value=float(m.avg_cost))

            # Color by margin
            margin_cell = ws.cell(row=idx, column=7)
            if m.margin_percent >= 50:
                margin_cell.fill = self.ABC_A_FILL
            elif m.margin_percent >= 30:
                margin_cell.fill = self.ABC_B_FILL
            else:
                margin_cell.fill = self.ABC_C_FILL

            for col in range(1, 10):
                self._apply_cell_style(ws.cell(row=idx, column=col))

        self._auto_column_width(ws)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
